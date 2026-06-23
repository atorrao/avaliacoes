"""
Vercel serverless function — generates ABANCA Excel report.
POST /api/generate-report
Body: { property: {...}, photos: [...], comps: [...] }
"""
from http.server import BaseHTTPRequestHandler
import json, os, sys, tempfile, urllib.request, base64
from datetime import datetime, date
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

TEMPLATE_URL = os.environ.get('REPORT_TEMPLATE_URL', '')

def v(val, default=''):
    return val if val is not None else default

def fmt_date(val):
    if not val: return ''
    if isinstance(val, (datetime, date)): return val.strftime('%d/%m/%Y')
    try: return datetime.strptime(str(val)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
    except: return str(val)

def fmt_area(val):
    if val is None: return ''
    try:
        f = float(val)
        if f == int(f): return str(int(f))
        return f'{f:.2f}'.replace('.', ',')
    except: return str(val)

def generate_report(data):
    p      = data.get('property', {})
    photos = data.get('photos', [])
    comps  = data.get('comps', [])

    # Download template
    tmpl_path = tempfile.mktemp(suffix='.xlsx')
    if TEMPLATE_URL:
        urllib.request.urlretrieve(TEMPLATE_URL, tmpl_path)
    else:
        # Fallback: use local copy if available
        local = '/var/task/template/Modelo_Relatorio_Abanca.xlsx'
        if os.path.exists(local):
            import shutil; shutil.copy(local, tmpl_path)
        else:
            raise FileNotFoundError('Template not found. Set REPORT_TEMPLATE_URL env var.')

    wb = load_workbook(tmpl_path)
    ws = wb['RELATÓRIO - PT']

    # 1. IDENTIFICAÇÃO
    ws['F8']  = fmt_date(p.get('data_relatorio') or datetime.today())
    ws['F10'] = v(p.get('nr_relatorio'), v(p.get('ref')))
    ws['X8']  = 'Abanca'
    ws['X9']  = v(p.get('tipo_servico'), 'Avaliação')
    ws['X10'] = v(p.get('finalidade'), 'Adjudicado sem visita interior')
    ws['X11'] = v(p.get('external_ref'), v(p.get('ref')))

    # 2. MORADA
    ws['D19'] = v(p.get('tipo_via'))
    ws['I19'] = v(p.get('street'), v(p.get('address')))
    ws['X19'] = v(p.get('number'))
    ws['Z19'] = v(p.get('floor_letter'))
    ws['AB19'] = v(p.get('fracao'))
    ws['AD19'] = v(p.get('block'))
    ws['D25'] = v(p.get('postal_code'))
    ws['I25'] = v(p.get('district'))
    ws['P25'] = v(p.get('municipality'))
    ws['W25'] = v(p.get('parish'))
    if p.get('longitude'): ws['D31'] = p['longitude']
    if p.get('latitude'):  ws['G31'] = p['latitude']

    # 3. DESCRIÇÃO
    ws['D38']  = v(p.get('property_type'))
    ws['K38']  = v(p.get('property_subtype'))
    ws['U38']  = v(p.get('use_type'))
    ws['AD38'] = v(p.get('use_subtype'))
    ws['D44']  = v(p.get('estado_construcao'), v(p.get('property_state')))
    ws['O44']  = v(p.get('destino'))
    ws['V44']  = v(p.get('estado_conservacao'))
    ws['AC44'] = v(p.get('estado_ocupacao'))
    ws['D50']  = v(p.get('composicao_imovel'), v(p.get('typology')))
    ws['D56']  = v(p.get('id_registo_predial'))
    ws['D62']  = v(p.get('id_registo_matricial'))
    ws['G62']  = v(p.get('fracao'))
    ws['D68']  = v(p.get('tipo_predio'))

    # 4. LOCALIZAÇÃO
    ws['J75'] = v(p.get('caract_mercado'))
    ws['J78'] = v(p.get('tipo_expectativa_mercado'))
    ws['J79'] = v(p.get('ocupacao_laboral'))
    ws['J80'] = v(p.get('populacao_concelho'))
    ws['J81'] = v(p.get('evolucao_mercado'), 'Tendencialmente positiva')

    # 5. CONSTRUÇÃO
    if p.get('nr_quartos'):          ws['D86'] = p['nr_quartos']
    if p.get('nr_inst_sanitarias'):  ws['G86'] = p['nr_inst_sanitarias']
    ws['J86'] = v(p.get('nr_pisos'), 1)
    ws['L86'] = v(p.get('qualidade_construcao'), 'Média')
    ws['P86'] = v(p.get('orientacao_solar'), 'Não influi no valor')
    ws['D92'] = v(p.get('nr_certificado_energ'))
    ws['J92'] = v(p.get('classe_energetica'))
    ws['N92'] = fmt_date(p.get('data_emissao_cert'))
    ws['R92'] = fmt_date(p.get('data_validade_cert'))
    ws['M98'] = v(p.get('year_built'))
    ws['D98'] = v(p.get('ano_licenca_utilizacao'))

    # 6. ÁREAS
    area_val = p.get('area_considerada') or p.get('area_m2') or p.get('gross_area')
    ws['D105'] = v(p.get('composicao_imovel'), v(p.get('typology')))
    ws['L105'] = fmt_area(p.get('land_area'))
    ws['Q105'] = fmt_area(area_val)
    ws['T105'] = fmt_area(p.get('area_annex_m2'))

    # 7. COMPARÁVEIS
    for idx, c in enumerate(comps[:3]):
        row = 116 + idx
        desc = v(c.get('notes'), f"{v(c.get('portal'))} ref.{v(c.get('listing_ref'))}")
        ws[f'D{row}'] = desc
        ws[f'T{row}'] = fmt_area(c.get('area_m2'))
        try:
            price = float(c.get('price') or 0)
            a     = float(c.get('area_m2') or 0)
            if a > 0:
                ws[f'Z{row}']  = round(price / a, 2)
                ws[f'AE{row}'] = price
        except: pass

    # 14. CONDICIONALISMOS
    ws['B248'] = v(p.get('prev_valuation_conditions'), 'Nenhum')

    # 16. CONCLUSÃO
    if p.get('valor_mercado'):          ws['D265'] = p['valor_mercado']
    if p.get('valor_venda_rapida'):     ws['J265'] = p['valor_venda_rapida']
    if p.get('valor_seguro'):           ws['R265'] = p['valor_seguro']
    if p.get('valor_mercado_atual'):    ws['D272'] = p['valor_mercado_atual']
    if p.get('valor_venda_rapida_atual'): ws['J272'] = p['valor_venda_rapida_atual']

    # 18. CERTIFICAÇÃO
    ws['K303'] = fmt_date(p.get('data_pedido_relatorio') or p.get('data_pedido'))
    ws['O303'] = fmt_date(p.get('data_visita') or p.get('visit_date'))
    ws['V303'] = fmt_date(p.get('data_conclusao') or p.get('data_relatorio'))
    ws['AC303'] = fmt_date(p.get('prev_valuation_date'))
    ws['D306']  = v(p.get('perito_avaliador'))

    # FOTOS
    if photos:
        sheet_name = 'Fotos do Imóvel'
        if sheet_name not in wb.sheetnames:
            ws_f = wb.create_sheet(sheet_name)
        else:
            ws_f = wb[sheet_name]

        ws_f['A1'] = 'RELATÓRIO DE AVALIAÇÃO IMOBILIÁRIA'
        ws_f['A1'].font = Font(bold=True, size=14)
        ws_f['A2'] = 'ANEXO — FOTOS DO IMÓVEL'
        ws_f['A2'].font = Font(bold=True, size=12)
        addr = v(p.get('street'), v(p.get('address')))
        ws_f['A3'] = f"Ref: {v(p.get('ref'))}  —  {addr}  —  {v(p.get('municipality'))}"
        ws_f.column_dimensions['A'].width = 45
        ws_f.column_dimensions['E'].width = 45

        row = 5
        col = 1
        for i, photo in enumerate(photos[:10]):
            url = photo.get('url')
            if not url: continue
            try:
                tmp = tempfile.mktemp(suffix='.jpg')
                urllib.request.urlretrieve(url, tmp)
                img = XLImage(tmp)
                img.width, img.height = 300, 220
                cell = f'{get_column_letter(col)}{row}'
                ws_f.add_image(img, cell)
                ws_f.row_dimensions[row].height = 165
                label_row = row + 15
                ws_f[f'{get_column_letter(col)}{label_row}'] = f'Foto {i + 1}'
                if col == 1:
                    col = 5
                else:
                    col = 1
                    row += 17
            except:
                ws_f[f'A{row}'] = f'[Foto {i+1} — indisponível]'
                row += 2

    # Save to temp
    out = tempfile.mktemp(suffix='.xlsx')
    wb.save(out)
    os.unlink(tmpl_path)
    with open(out, 'rb') as f: content = f.read()
    os.unlink(out)
    return content


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length  = int(self.headers.get('Content-Length', 0))
        body    = self.rfile.read(length)
        data    = json.loads(body)
        try:
            xlsx = generate_report(data)
            ref  = data.get('property', {}).get('ref', 'relatorio')
            name = f'Relatorio_{ref}_{datetime.today().strftime("%Y%m%d")}.xlsx'
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{name}"')
            self.send_header('Content-Length', len(xlsx))
            self.end_headers()
            self.wfile.write(xlsx)
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())
